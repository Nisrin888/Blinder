from __future__ import annotations

import csv
import io
import logging
from collections import Counter
from uuid import UUID

from sqlalchemy.ext.asyncio import AsyncSession

from blinder.encryption import derive_key
from blinder.pii_detector import PIIDetector, PIIEntity
from blinder.pipeline import BlinderPipeline
from blinder.vault import Vault, VaultEntry
from config import get_settings
from db import repositories
from schemas.api import DocumentResponse
from services.embedding_service import EmbeddingService

SEPARATOR = " | "
SEP_LEN = len(SEPARATOR)  # 3
SAMPLE_SIZE = 5  # data rows to sample per column for PII detection

logger = logging.getLogger(__name__)
settings = get_settings()


async def extract_text(file_content: bytes, content_type: str) -> str:
    """Extract plain text from an uploaded file.

    Supports PDF (via pypdf), DOCX (via python-docx), and plain text.
    """
    if content_type == "application/pdf":
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(file_content))
        pages = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
        return "\n\n".join(pages)

    elif content_type in (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    ):
        from docx import Document

        doc = Document(io.BytesIO(file_content))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)

    elif content_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        sheets: list[str] = []
        for ws in wb.worksheets:
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append(" | ".join(cells))
            if rows:
                header = f"[Sheet: {ws.title}]"
                sheets.append(f"{header}\n{rows[0]}\n" + "\n".join(rows[1:]))
        wb.close()
        return "\n\n".join(sheets)

    elif content_type == "text/csv":
        text = file_content.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = [" | ".join(row) for row in reader if any(row)]
        return "\n".join(rows)

    else:
        # Treat as plain text
        return file_content.decode("utf-8", errors="replace")


def _parse_tabular_rows(file_content: bytes, content_type: str) -> list[list[str]]:
    """Parse raw file bytes into a list of rows (list of cell strings).

    Supports CSV and Excel formats.
    """
    if content_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        rows: list[list[str]] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append(cells)
        wb.close()
        return rows

    # Default: CSV / TSV
    text = file_content.decode("utf-8", errors="replace")
    return [row for row in csv.reader(io.StringIO(text)) if any(row)]


async def _detect_pii_columns(
    rows: list[list[str]],
    detector: PIIDetector,
    sample_size: int = SAMPLE_SIZE,
) -> dict[int, str]:
    """Run NER on sample cell values per column to identify PII columns.

    Returns {column_index: entity_type} for columns containing PII.
    """
    if len(rows) < 2:  # need at least header + 1 data row
        return {}

    headers = rows[0]
    data_rows = rows[1 : 1 + sample_size]
    pii_columns: dict[int, str] = {}

    for col_idx in range(len(headers)):
        sample_values = [
            row[col_idx]
            for row in data_rows
            if col_idx < len(row) and row[col_idx].strip()
        ]
        if not sample_values:
            continue

        # Concatenate with newlines — run full detection (both gates)
        sample_text = "\n".join(sample_values)
        entities = await detector.detect(sample_text, skip_ner=False)

        if entities:
            type_counts = Counter(e.label for e in entities)
            dominant_type = type_counts.most_common(1)[0][0]
            pii_columns[col_idx] = dominant_type

    return pii_columns


def _build_column_entities(
    full_text: str,
    rows: list[list[str]],
    pii_columns: dict[int, str],
) -> list[PIIEntity]:
    """Generate PIIEntity objects for every cell in identified PII columns.

    Computes character offsets by walking through the pipe-delimited text
    row by row, column by column.
    """
    entities: list[PIIEntity] = []
    text_offset = 0

    for row_idx, row in enumerate(rows):
        line = SEPARATOR.join(row)

        if row_idx == 0:
            # Skip header row — column names aren't PII
            text_offset += len(line) + 1  # +1 for \n
            continue

        col_offset = text_offset
        for col_idx, cell_value in enumerate(row):
            if col_idx in pii_columns and cell_value.strip():
                entities.append(PIIEntity(
                    text=cell_value,
                    label=pii_columns[col_idx],
                    start=col_offset,
                    end=col_offset + len(cell_value),
                    confidence=0.90,
                    gate="column",
                ))
            col_offset += len(cell_value)
            if col_idx < len(row) - 1:
                col_offset += SEP_LEN

        text_offset += len(line) + 1  # +1 for \n

    return entities


def _is_tabular(text: str) -> bool:
    """Detect if text is pipe-delimited tabular data (CSV/Excel output)."""
    lines = text.split("\n", 5)  # check first few lines
    pipe_lines = sum(1 for line in lines if line.count(" | ") >= 2)
    return pipe_lines >= 2


def _chunk_text(text: str, chunk_size: int = 512, overlap: int = 50) -> list[str]:
    """Split text into chunks. For tabular data, chunks by rows with header
    prepended to every chunk so the LLM always knows what each column means."""
    if not text.strip():
        return []

    # Tabular data: chunk by rows, prepend header to each chunk
    if _is_tabular(text):
        return _chunk_tabular(text, chunk_size)

    # Prose: word-based chunks with overlap
    words = text.split()
    if len(words) <= chunk_size:
        return [text]

    chunks = []
    start = 0
    while start < len(words):
        end = start + chunk_size
        chunk = " ".join(words[start:end])
        chunks.append(chunk)
        start = end - overlap
    return chunks


def _chunk_tabular(text: str, chunk_size: int = 512) -> list[str]:
    """Chunk tabular (pipe-delimited) text by rows, prepending the header row
    to every chunk so the LLM can always identify columns."""
    lines = text.split("\n")
    if len(lines) < 2:
        return [text] if text.strip() else []

    header = lines[0]
    header_words = len(header.split())
    data_lines = lines[1:]

    # Budget per chunk: chunk_size minus header overhead
    budget = max(chunk_size - header_words, chunk_size // 2)

    chunks = []
    current_lines: list[str] = []
    current_words = 0

    for line in data_lines:
        if not line.strip():
            continue
        line_words = len(line.split())
        if current_words + line_words > budget and current_lines:
            # Emit chunk with header prepended
            chunks.append(header + "\n" + "\n".join(current_lines))
            current_lines = []
            current_words = 0
        current_lines.append(line)
        current_words += line_words

    # Final chunk
    if current_lines:
        chunks.append(header + "\n" + "\n".join(current_lines))

    return chunks


async def process_document(
    db: AsyncSession,
    session_id: UUID,
    filename: str,
    content_type: str,
    file_content: bytes,
) -> tuple[DocumentResponse, dict[str, int], list]:
    """Upload and process a document through the blinding pipeline.

    Returns
    -------
    tuple of (DocumentResponse, pii_summary, threats)
        pii_summary maps entity type to count of detections.
    """
    # 1. Extract text from the uploaded file
    text = await extract_text(file_content, content_type)

    # 2. Load or create vault for the session
    session = await repositories.get_session(db, session_id)
    if session is None:
        raise ValueError(f"Session {session_id} not found")

    encryption_key = derive_key(settings.blinder_master_key, session.session_salt)
    vault = Vault(session_salt=session.session_salt, encryption_key=encryption_key)

    # Load existing vault entries from DB
    db_entries = await repositories.get_vault_entries(db, session_id)
    if db_entries:
        loaded_entries = []
        for entry in db_entries:
            real_value = vault.decrypt_value(entry.encrypted_value, entry.nonce)
            loaded_entries.append(
                VaultEntry(
                    entity_type=entry.entity_type,
                    pseudonym=entry.pseudonym,
                    real_value=real_value,
                    aliases=entry.aliases or [],
                )
            )
        vault.load_entries(loaded_entries)

    # 3. Create BlinderPipeline and process the document
    #    For tabular formats, use sample-based column detection: run NER on
    #    a small sample of each column to identify PII columns, then mask
    #    every cell in those columns without running NER on the full file.
    tabular_types = {
        "text/csv",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    }
    tabular_extensions = {".csv", ".xlsx", ".xls", ".tsv"}
    file_ext = ("." + filename.rsplit(".", 1)[-1]).lower() if "." in filename else ""
    is_tabular = content_type in tabular_types or file_ext in tabular_extensions

    pipeline = BlinderPipeline(vault)

    if is_tabular:
        logger.info(
            "Tabular format detected (type=%s, ext=%s) — using sample-based column detection",
            content_type, file_ext,
        )

        # Parse tabular rows from raw bytes (not from pipe-delimited text)
        parsed_rows = _parse_tabular_rows(file_content, content_type)

        # Rebuild pipe-delimited text from parsed rows so offsets are consistent
        text = "\n".join(SEPARATOR.join(row) for row in parsed_rows)

        # Detect which columns contain PII by sampling a few rows
        pii_columns = await _detect_pii_columns(parsed_rows, pipeline._detector)

        if pii_columns:
            logger.info("PII columns detected: %s", {
                parsed_rows[0][i] if i < len(parsed_rows[0]) else f"col_{i}": t
                for i, t in pii_columns.items()
            })
            # Generate entities for ALL cells in PII columns
            column_entities = _build_column_entities(text, parsed_rows, pii_columns)
        else:
            column_entities = []

        # Also run pattern-only Presidio on full text (catches SSNs, IPs, etc.)
        pattern_entities = await pipeline._detector.detect(text, skip_ner=True)

        # Merge both sets, dedup overlapping spans
        all_entities = PIIDetector._merge_detections(column_entities, pattern_entities)

        # Process with pre-detected entities (skips PII detection, still runs threat sanitiser)
        blinded_text, pii_count, threats = await pipeline.process_document_with_entities(
            text, all_entities
        )
    else:
        logger.info(
            "Prose format detected (type=%s, ext=%s) — running both gates",
            content_type, file_ext,
        )
        blinded_text, pii_count, threats = await pipeline.process_document(text)

    # 4. Save the document to DB
    doc = await repositories.create_document(
        db,
        session_id=session_id,
        filename=filename,
        content_type=content_type,
        raw_text=text,
    )
    doc = await repositories.update_document_processed(
        db,
        doc_id=doc.id,
        blinded_text=blinded_text,
        pii_count=pii_count,
    )

    # 5. Chunk + embed for hybrid RAG (prose documents only)
    #    Tabular data (CSV/XLS) is queried directly via tabular_query.py
    #    from blinded_text — no chunking or embedding needed.
    if not is_tabular:
        chunks = _chunk_text(blinded_text, settings.chunk_size, settings.chunk_overlap)
        if chunks:
            embedder = EmbeddingService()
            embeddings = embedder.embed_batch(chunks)
            chunk_records = [
                {
                    "session_id": session_id,
                    "document_id": doc.id,
                    "chunk_index": i,
                    "content": chunk,
                    "embedding": embedding,
                }
                for i, (chunk, embedding) in enumerate(zip(chunks, embeddings))
            ]
            await repositories.create_chunks_bulk(db, chunk_records)
            logger.info("Created %d chunks for document %s", len(chunks), doc.id)
    else:
        logger.info("Tabular document — skipping RAG chunking (queried directly via tabular_query)")

    # 6. Save all new vault entries to DB
    existing_pseudonyms = {e.pseudonym for e in db_entries}
    for vault_entry in vault.get_all_entries():
        if vault_entry.pseudonym not in existing_pseudonyms:
            encrypted_value, nonce = vault.encrypt_value(vault_entry.real_value)
            await repositories.create_vault_entry(
                db,
                session_id=session_id,
                entity_type=vault_entry.entity_type,
                pseudonym=vault_entry.pseudonym,
                encrypted_value=encrypted_value,
                nonce=nonce,
                aliases=vault_entry.aliases,
            )

    # 7. Build PII summary (count by entity type)
    pii_summary: dict[str, int] = dict(
        Counter(entry.entity_type for entry in vault.get_all_entries())
    )

    # 8. Build response
    doc_response = DocumentResponse.model_validate(doc)

    threat_dicts = [
        {
            "threat_type": t.threat_type,
            "description": t.description,
            "severity": t.severity,
        }
        for t in threats
    ]

    return doc_response, pii_summary, threat_dicts
